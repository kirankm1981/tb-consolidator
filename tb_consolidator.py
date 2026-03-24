import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def determine_category(cell):
    font = cell.font
    fill = cell.fill

    is_bold = font.bold is True
    is_italic = font.italic is True
    has_fill = fill.patternType is not None and fill.patternType != "none"

    if has_fill:
        return "Group Accounts"
    if is_italic:
        return "Sub-GL Accounts"
    if is_bold:
        return "Control Accounts"
    return "GL Accounts"


def extract_entity_name(ws):
    cell = ws.cell(row=1, column=1)
    raw_value = cell.value
    if raw_value is None:
        return ""
    text = str(raw_value)
    first_line = text.split("\n")[0].strip()
    return first_line


def find_last_data_row(ws):
    max_row = ws.max_row
    for r in range(max_row, 6, -1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            text = str(val).strip().lower()
            if text.startswith("grand total") or text == "total" or text == "totals":
                return r - 1
    return max_row


def to_num(val):
    if isinstance(val, (int, float)):
        return val
    if val is None or val == "":
        return 0
    try:
        return float(str(val).replace(",", ""))
    except ValueError:
        return 0


def load_mapping_file(filepath):
    company_codes = {}
    fsli_codes = {}

    if not filepath:
        return company_codes, fsli_codes

    wb = load_workbook(filepath, data_only=True)

    if "Company_Code" in wb.sheetnames:
        ws = wb["Company_Code"]
        for r in range(2, ws.max_row + 1):
            erp_name = ws.cell(r, 2).value
            code = ws.cell(r, 3).value
            if erp_name and code:
                company_codes[str(erp_name).strip().lower()] = str(code).strip()

    if "FSLI_Code" in wb.sheetnames:
        ws = wb["FSLI_Code"]
        for r in range(2, ws.max_row + 1):
            gl_code = str(ws.cell(r, 1).value or "").strip()
            gl_desc = str(ws.cell(r, 2).value or "").strip()
            fs_header = str(ws.cell(r, 3).value or "").strip()
            fs_account_type = str(ws.cell(r, 4).value or "").strip()
            fs_account_sub_type = str(ws.cell(r, 5).value or "").strip()
            fsli = str(ws.cell(r, 6).value or "").strip()
            if gl_desc:
                fsli_codes[gl_desc.lower()] = {
                    "gl_code": gl_code,
                    "fs_header": fs_header,
                    "fs_account_type": fs_account_type,
                    "fs_account_sub_type": fs_account_sub_type,
                    "fsli": fsli,
                }

    return company_codes, fsli_codes


def process_file(filepath, company_codes, fsli_codes):
    rows = []
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    entity_name = extract_entity_name(ws)
    if not entity_name:
        return rows, f"Could not extract entity name from {os.path.basename(filepath)}"

    entity_code = company_codes.get(entity_name.lower(), "")
    last_data_row = find_last_data_row(ws)

    for r in range(7, last_data_row + 1):
        cell_a = ws.cell(row=r, column=1)
        account_head = cell_a.value
        if account_head is None or str(account_head).strip() == "":
            continue

        category = determine_category(cell_a)
        account_head_str = str(account_head).strip()

        def get_value(col):
            val = ws.cell(row=r, column=col).value
            if val is None or val == "":
                return ""
            if isinstance(val, (int, float)):
                return val
            text = str(val).replace(",", "")
            try:
                return float(text)
            except ValueError:
                return val

        opening_debit = get_value(2)
        opening_credit = get_value(3)
        closing_debit = get_value(6)
        closing_credit = get_value(7)

        opening_balance = to_num(opening_debit) - to_num(opening_credit)
        closing_balance = to_num(closing_debit) - to_num(closing_credit)

        fsli_entry = fsli_codes.get(account_head_str.lower(), {})

        rows.append({
            "entity_name": entity_name,
            "entity_code": entity_code,
            "category": category,
            "account_head": account_head_str,
            "company_gl_code": fsli_entry.get("gl_code", ""),
            "opening_debit": opening_debit,
            "opening_credit": opening_credit,
            "opening_balance": opening_balance,
            "period_debit": get_value(4),
            "period_credit": get_value(5),
            "closing_debit": closing_debit,
            "closing_credit": closing_credit,
            "closing_balance": closing_balance,
            "fs_header": fsli_entry.get("fs_header", ""),
            "fs_account_type": fsli_entry.get("fs_account_type", ""),
            "fs_account_sub_type": fsli_entry.get("fs_account_sub_type", ""),
            "fsli": fsli_entry.get("fsli", ""),
        })

    return rows, None


def consolidate_files(folder_path, output_path, mapping_path=None, progress_callback=None):
    xlsx_files = [
        os.path.join(folder_path, f)
        for f in os.listdir(folder_path)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not xlsx_files:
        return 0, [], ["No .xlsx files found in the selected folder"]

    company_codes, fsli_codes = load_mapping_file(mapping_path)

    all_rows = []
    processed_files = []
    errors = []

    for i, filepath in enumerate(xlsx_files):
        filename = os.path.basename(filepath)
        if progress_callback:
            progress_callback(i + 1, len(xlsx_files), filename)
        try:
            rows, error = process_file(filepath, company_codes, fsli_codes)
            if error:
                errors.append(error)
            else:
                all_rows.extend(rows)
                processed_files.append(filename)
        except Exception as e:
            errors.append(f"{filename}: {str(e)}")

    if not all_rows:
        return 0, processed_files, errors

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Trial Balance"

    headers = [
        "Period",
        "Company GL Code",
        "Company GL Description",
        "Company Name",
        "Company Code",
        "Opening Balance",
        "Debit Amount",
        "Credit Amount",
        "Closing Balance",
        "Currency",
        "Category",
        "FS Header",
        "FS Account Type",
        "FS Account Sub-Type",
        "FSLI",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        top=Side(style="thin"),
        bottom=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )
    light_border = Border(
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(all_rows, 2):
        values = [
            "",
            row_data["company_gl_code"],
            row_data["account_head"],
            row_data["entity_name"],
            row_data["entity_code"],
            row_data["opening_balance"],
            row_data["period_debit"],
            row_data["period_credit"],
            row_data["closing_balance"],
            "INR",
            row_data["category"],
            row_data["fs_header"],
            row_data["fs_account_type"],
            row_data["fs_account_sub_type"],
            row_data["fsli"],
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.border = light_border
            if 6 <= col_num <= 9 and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    col_widths = [15, 18, 50, 40, 18, 18, 18, 18, 18, 12, 20, 20, 20, 25, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = f"A1:O{len(all_rows) + 1}"

    cs = wb.create_sheet("Control")

    control_filtered = [r for r in all_rows if r["category"] in ("GL Accounts", "Control Accounts")]

    control_summary = {}
    for row_data in control_filtered:
        key = row_data["entity_code"] or row_data["entity_name"]
        if key in control_summary:
            control_summary[key]["closing_balance"] += to_num(row_data["closing_balance"])
        else:
            control_summary[key] = {
                "company_name": row_data["entity_name"],
                "closing_balance": to_num(row_data["closing_balance"]),
            }

    ctrl_headers = ["Company Code", "Company Name-ERP", "Closing Balance"]
    for col_num, header in enumerate(ctrl_headers, 1):
        cell = cs.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, (code, data) in enumerate(control_summary.items(), 2):
        cs.cell(row=idx, column=1, value=code).border = light_border
        cs.cell(row=idx, column=2, value=data["company_name"]).border = light_border
        cell = cs.cell(row=idx, column=3, value=data["closing_balance"])
        cell.border = light_border
        cell.number_format = "#,##0.00"

    cs.column_dimensions["A"].width = 20
    cs.column_dimensions["B"].width = 45
    cs.column_dimensions["C"].width = 22
    cs.auto_filter.ref = f"A1:C{len(control_summary) + 1}"

    wb.save(output_path)
    return len(all_rows), processed_files, errors


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Trial Balance Consolidator")
        self.root.geometry("700x550")
        self.root.resizable(True, True)

        style = ttk.Style()
        style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"))
        style.configure("Info.TLabel", font=("Segoe UI", 10))
        style.configure("Action.TButton", font=("Segoe UI", 11))

        main_frame = ttk.Frame(root, padding=30)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="Trial Balance Consolidator", style="Title.TLabel").pack(pady=(0, 5))
        ttk.Label(
            main_frame,
            text="Select a folder containing Trial Balance Excel files to consolidate them.",
            style="Info.TLabel",
        ).pack(pady=(0, 20))

        folder_frame = ttk.Frame(main_frame)
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(folder_frame, text="Source Folder:").pack(side=tk.LEFT)
        self.folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_var, state="readonly", width=50).pack(
            side=tk.LEFT, padx=(10, 10), fill=tk.X, expand=True
        )
        ttk.Button(folder_frame, text="Browse", command=self.select_folder).pack(side=tk.LEFT)

        mapping_frame = ttk.Frame(main_frame)
        mapping_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(mapping_frame, text="Mapping File:").pack(side=tk.LEFT)
        self.mapping_var = tk.StringVar(value="(Optional)")
        ttk.Entry(mapping_frame, textvariable=self.mapping_var, state="readonly", width=50).pack(
            side=tk.LEFT, padx=(10, 10), fill=tk.X, expand=True
        )
        ttk.Button(mapping_frame, text="Browse", command=self.select_mapping).pack(side=tk.LEFT)

        self.mapping_path = None

        self.consolidate_btn = ttk.Button(
            main_frame, text="Consolidate & Save", command=self.consolidate, style="Action.TButton"
        )
        self.consolidate_btn.pack(pady=20)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))

        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(main_frame, textvariable=self.status_var, style="Info.TLabel").pack()

        self.log_text = tk.Text(main_frame, height=10, font=("Consolas", 9), state="disabled")
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        scrollbar = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)

    def log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        self.root.update_idletasks()

    def select_folder(self):
        folder = filedialog.askdirectory(title="Select folder containing Trial Balance files")
        if folder:
            self.folder_var.set(folder)
            xlsx_count = len([f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")])
            self.log(f"Selected folder: {folder}")
            self.log(f"Found {xlsx_count} Excel file(s)")

    def select_mapping(self):
        filepath = filedialog.askopenfilename(
            title="Select Mapping File",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if filepath:
            self.mapping_path = filepath
            self.mapping_var.set(os.path.basename(filepath))
            self.log(f"Mapping file: {os.path.basename(filepath)}")

    def update_progress(self, current, total, filename):
        self.progress_var.set((current / total) * 100)
        self.status_var.set(f"Processing {current}/{total}: {filename}")
        self.root.update_idletasks()

    def consolidate(self):
        folder = self.folder_var.get()
        if not folder:
            messagebox.showwarning("No Folder", "Please select a folder first.")
            return

        output_path = filedialog.asksaveasfilename(
            title="Save consolidated file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Consolidated_Trial_Balance.xlsx",
        )
        if not output_path:
            return

        self.consolidate_btn.config(state="disabled")
        self.log("\nStarting consolidation...")

        try:
            total_rows, processed_files, errors = consolidate_files(
                folder, output_path,
                mapping_path=self.mapping_path,
                progress_callback=self.update_progress,
            )

            self.log(f"\nProcessed {len(processed_files)} file(s) successfully:")
            for f in processed_files:
                self.log(f"  - {f}")

            if errors:
                self.log(f"\nWarnings ({len(errors)}):")
                for e in errors:
                    self.log(f"  ! {e}")

            self.log(f"\nTotal rows consolidated: {total_rows:,}")
            self.log(f"Output saved to: {output_path}")

            self.progress_var.set(100)
            self.status_var.set("Consolidation complete!")
            messagebox.showinfo(
                "Complete",
                f"Consolidated {len(processed_files)} file(s) with {total_rows:,} rows.\n\nSaved to:\n{output_path}",
            )
        except Exception as e:
            self.log(f"\nError: {str(e)}")
            self.status_var.set("Error occurred")
            messagebox.showerror("Error", str(e))
        finally:
            self.consolidate_btn.config(state="normal")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
